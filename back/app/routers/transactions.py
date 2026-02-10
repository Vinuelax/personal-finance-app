from typing import List, Optional, Tuple
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel, ConfigDict, Field
from datetime import datetime, timezone, date
from io import BytesIO
import csv
import re

from utils.deps import get_db, get_current_user
from utils.db import DB, list_transactions, get_transaction, create_transaction, delete_transaction
from copy import deepcopy

router = APIRouter(tags=["transactions"])

SPANISH_MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


class TransactionSplit(BaseModel):
    id: str
    label: str
    amount: int
    categoryId: Optional[str] = None


class TransactionIn(BaseModel):
    model_config = ConfigDict(json_schema_extra={
        "example": {
            "date": "2026-02-05",
            "merchant": "Starbucks",
            "description": "Latte and croissant",
            "amount": -6200,
            "currency": "CLP",
            "categoryId": "cat_coffee",
            "notes": "Morning treat",
            "source": "manual",
            "accountId": "acc_bank_001",
            "receiptId": None,
        }
    })

    date: str = Field(..., description="Posting date in YYYY-MM-DD")
    merchant: str = Field(..., description="Merchant or payee name")
    description: Optional[str] = Field("", description="Free-form description")
    amount: int = Field(..., description="Signed amount in minor units (negative for expenses)")
    currency: str = Field("CLP", description="ISO currency code")
    categoryId: Optional[str] = Field(None, description="Category ID, if already assigned")
    notes: Optional[str] = Field("", description="User notes")
    source: str = Field("manual", description="Origin of the transaction, e.g. manual, bank_scrape")
    accountId: Optional[str] = Field(None, description="Account identifier")
    receiptId: Optional[str] = Field(None, description="Linked receipt ID")
    splits: Optional[List[TransactionSplit]] = Field(None, description="Optional list of splits with amount, label, categoryId")


class TransactionUpdate(BaseModel):
    model_config = ConfigDict(json_schema_extra={
        "example": {
            "description": "Updated memo",
            "categoryId": "cat_groceries",
            "notes": "Matched to weekly shop"
        }
    })

    merchant: Optional[str] = Field(None, description="Merchant or payee name")
    description: Optional[str] = Field(None, description="Free-form description")
    amount: Optional[int] = Field(None, description="Signed amount in minor units")
    currency: Optional[str] = Field(None, description="ISO currency code")
    categoryId: Optional[str] = Field(None, description="Category ID")
    notes: Optional[str] = Field(None, description="User notes")
    source: Optional[str] = Field(None, description="Origin of the transaction")
    accountId: Optional[str] = Field(None, description="Account identifier")
    receiptId: Optional[str] = Field(None, description="Linked receipt ID")
    splits: Optional[List[TransactionSplit]] = Field(None, description="Optional list of splits to replace existing")


class TransactionOut(TransactionIn):
    model_config = ConfigDict(json_schema_extra={
        "example": {
            "txnId": "txn_000123",
            "date": "2026-02-05",
            "merchant": "Starbucks",
            "description": "Latte and croissant",
            "amount": -6200,
            "currency": "CLP",
            "categoryId": "cat_coffee",
            "notes": "Morning treat",
            "source": "manual",
            "accountId": "acc_bank_001",
            "receiptId": None,
        }
    })

    txnId: str = Field(..., description="Transaction identifier")


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _has_split_category(item: dict) -> bool:
    splits = item.get("splits") or []
    return any(s.get("categoryId") for s in splits)


def _is_uncategorized(item: dict) -> bool:
    """Uncategorized = no main category and no split category, and expense (amount < 0)."""
    return (item.get("amount", 0) < 0) and (not item.get("categoryId")) and (not _has_split_category(item))


def _apply_uncat_index(item: dict, user_id: str):
    """Mutates item to add or remove GSI2 keys based on uncategorized status."""
    if _is_uncategorized(item):
        item["GSI2PK"] = f"USER#{user_id}#UNCAT"
        item["GSI2SK"] = f"DATE#{item.get('date')}#TX#{item.get('txnId')}"
    else:
        item.pop("GSI2PK", None)
        item.pop("GSI2SK", None)


def _parse_date_like(value: str, fallback_year: int | None = None, statement_end: date | None = None) -> Optional[str]:
    """
    Accepts dd/mm or dd/mm/yyyy or ISO date; returns YYYY-MM-DD.
    If only dd/mm is provided, infer year using statement_end (preferred) or fallback_year.
    If month > statement_end.month, assume previous year (to handle statements spanning year boundary).
    """
    if not value:
        return None
    v = str(value).strip()
    # ISO
    try:
        dt = datetime.fromisoformat(v.replace("Z", "+00:00"))
        return dt.date().isoformat()
    except Exception:
        pass
    # dd/mm or dd/mm/yyyy
    m = re.match(r"^(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?$", v)
    if m:
        d, mth, yr = m.groups()
        day = int(d)
        month = int(mth)
        year = None
        if yr:
            year = int(yr) + 2000 if len(yr) == 2 else int(yr)
        elif statement_end:
            year = statement_end.year
            # If statement ends in Dec and we see Jan without a year, assume next year (bank year rollover)
            if month < statement_end.month and statement_end.month == 12:
                year += 1
            # If month appears ahead of statement_end (e.g., end is Jan and value is Dec), assume previous year
            elif month > statement_end.month:
                year -= 1
        elif fallback_year:
            year = fallback_year
        if year:
            return date(year, month, day).isoformat()
    return None


def _detect_statement_dates(ws) -> Tuple[Optional[date], Optional[date]]:
    start = None
    end = None
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, datetime):
                d = cell.date()
                if not start or d < start:
                    start = d
                if not end or d > end:
                    end = d
            elif isinstance(cell, str):
                m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", cell.strip())
                if m:
                    d, mth, yr = m.groups()
                    year = int(yr) + 2000 if len(yr) == 2 else int(yr)
                    parsed = date(year, int(mth), int(d))
                    if not start or parsed < start:
                        start = parsed
                    if not end or parsed > end:
                        end = parsed
        if start and end:
            break
    return start, end


def _detect_year_hint(ws) -> Tuple[Optional[int], Optional[int]]:
    """
    Try to read the headline like 'Cartola de cuenta Corriente - Enero 2026'
    to get (year, month). We only scan top rows to keep it fast.
    """
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        for cell in row:
            if not isinstance(cell, str):
                continue
            m = re.search(r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre)\\s+(\\d{4})", cell, re.IGNORECASE)
            if m:
                month = SPANISH_MONTHS[m.group(1).lower()]
                year = int(m.group(2))
                return year, month
    return None, None


def _parse_excel_transactions(content: bytes) -> List[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    year_hint, month_hint = _detect_year_hint(ws)

    header_row_idx = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and any(str(c).strip().upper() == "FECHA" for c in row):
            header_row_idx = idx
            headers = [str(c).strip().upper() if c else "" for c in row]
            break
    if not header_row_idx:
        return []

    charges_idx = headers.index("CHEQUES Y OTROS CARGOS") if "CHEQUES Y OTROS CARGOS" in headers else None
    deposits_idx = headers.index("DEPOSITOS Y OTROS ABONOS") if "DEPOSITOS Y OTROS ABONOS" in headers else None
    desc_idx = headers.index("DESCRIPCIÓN") if "DESCRIPCIÓN" in headers else headers.index("DESCRIPCION") if "DESCRIPCION" in headers else None

    statement_start, statement_end = _detect_statement_dates(ws)
    fallback_year = statement_end.year if statement_end else year_hint

    transactions = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(row):
            continue
        date_raw = row[0]
        date_iso = None
        if isinstance(date_raw, (datetime, date)):
            date_iso = (date_raw.date() if isinstance(date_raw, datetime) else date_raw).isoformat()
        elif date_raw:
            date_iso = _parse_date_like(str(date_raw), fallback_year=fallback_year, statement_end=statement_end)
        if not date_iso:
            continue

        charge = row[charges_idx] if charges_idx is not None else None
        deposit = row[deposits_idx] if deposits_idx is not None else None
        amount = None
        if charge not in (None, "", 0):
            try:
                amount = -abs(float(charge))
            except Exception:
                pass
        if amount is None and deposit not in (None, "", 0):
            try:
                amount = abs(float(deposit))
            except Exception:
                pass
        if amount is None:
            continue

        description = row[desc_idx] if desc_idx is not None else ""
        transactions.append({
            "date": date_iso,
            "merchant": str(description or "Transaction")[:120],
            "description": str(description or ""),
            "amount": int(round(amount)),
            "currency": "CLP",
            "categoryId": None,
            "notes": "",
            "source": "upload",
            "accountId": None,
            "receiptId": None,
        })
    return transactions


def _parse_csv_transactions(content: bytes) -> List[dict]:
    text = content.decode("utf-8", errors="ignore")
    sample = text[:1024]
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(sample)
    except Exception:
        dialect = csv.excel
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    transactions = []
    for row in reader:
        date_raw = row.get("date") or row.get("fecha") or row.get("Date")
        date_iso = _parse_date_like(str(date_raw)) if date_raw else None
        if not date_iso:
            continue
        amount_raw = row.get("amount") or row.get("monto") or row.get("Amount")
        if amount_raw is None:
            continue
        try:
            amount_val = float(str(amount_raw).replace(",", ""))
        except Exception:
            continue
        amount_minor = int(round(amount_val * 100)) if abs(amount_val) < 100000 and not float(amount_val).is_integer() else int(round(amount_val))
        desc = row.get("description") or row.get("descripcion") or row.get("merchant") or ""
        transactions.append({
            "date": date_iso,
            "merchant": str(desc or "Transaction")[:120],
            "description": str(desc or ""),
            "amount": amount_minor,
            "currency": "CLP",
            "categoryId": None,
            "notes": "",
            "source": "upload",
            "accountId": None,
            "receiptId": None,
        })
    return transactions


@router.get(
    "",
    response_model=List[TransactionOut],
    summary="List transactions",
    description="Returns transactions for the authenticated user filtered by optional date range, category, uncategorized flag, or limit."
)
def api_list_transactions(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    category_id: Optional[str] = None,
    uncategorized: bool = False,
    limit: Optional[int] = None,
    current_user=Depends(get_current_user),
    db: DB = Depends(get_db),
):
    items = list_transactions(db, current_user["user_id"], date_from, date_to, category_id, uncategorized, limit)
    return [
        {
            "txnId": i.get("txnId"),
            "date": i.get("date"),
            "merchant": i.get("merchant"),
            "description": i.get("description"),
            "amount": i.get("amount"),
            "currency": i.get("currency"),
            "categoryId": i.get("categoryId"),
            "notes": i.get("notes", ""),
            "source": i.get("source"),
            "accountId": i.get("accountId"),
            "receiptId": i.get("receiptId"),
            "splits": i.get("splits"),
        }
        for i in items
    ]


@router.post(
    "",
    response_model=TransactionOut,
    summary="Create transaction",
    description="Create a manual transaction for the authenticated user."
)
def api_create_transaction(payload: TransactionIn, current_user=Depends(get_current_user), db: DB = Depends(get_db)):
    data = payload.model_dump()
    splits = data.get("splits") or []
    if splits:
        data["splits"] = [dict(s) for s in splits]
        if len(splits) > 1:
            data["categoryId"] = None
    created = create_transaction(db, current_user["user_id"], data)
    return {k: created.get(k) for k in ["txnId", "date", "merchant", "description", "amount", "currency", "categoryId", "notes", "source", "accountId", "receiptId", "splits"]}


@router.get(
    "/{txn_id}",
    response_model=TransactionOut,
    summary="Get transaction",
    description="Fetch a single transaction by ID and date."
)
def api_get_transaction(txn_id: str, date: str, current_user=Depends(get_current_user), db: DB = Depends(get_db)):
    item = get_transaction(db, current_user["user_id"], txn_id, date)
    if not item:
        raise HTTPException(404, "Transaction not found")
    return {k: item.get(k) for k in ["txnId", "date", "merchant", "description", "amount", "currency", "categoryId", "notes", "source", "accountId", "receiptId", "splits"]}


@router.patch(
    "/{txn_id}",
    response_model=TransactionOut,
    summary="Update transaction",
    description="Update mutable fields of a transaction."
)
def api_update_transaction(txn_id: str, payload: TransactionUpdate, date: str, current_user=Depends(get_current_user), db: DB = Depends(get_db)):
    updates = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    existing = db.get_transaction(current_user["user_id"], txn_id)
    if not existing:
        raise HTTPException(404, "Transaction not found")

    # If splits provided, clear main category when multiple parts
    splits = updates.get("splits") or existing.get("splits") or []
    if splits:
        updates["splits"] = [dict(s) for s in splits]
        if len(splits) > 1:
            updates["categoryId"] = None
    updated = db.update_transaction(current_user["user_id"], txn_id, updates)
    if not updated:
        raise HTTPException(500, "Failed to persist transaction")
    return {k: updated.get(k) for k in ["txnId", "date", "merchant", "description", "amount", "currency", "categoryId", "notes", "source", "accountId", "receiptId", "splits"]}


@router.delete(
    "/{txn_id}",
    summary="Delete transaction",
    description="Delete a transaction by ID and date."
)
def api_delete_transaction(txn_id: str, date: str, current_user=Depends(get_current_user), db: DB = Depends(get_db)):
    ok = delete_transaction(db, current_user["user_id"], txn_id, date)
    if not ok:
        raise HTTPException(404, "Transaction not found")
    return {"deleted": True}


@router.post(
    "/import",
    summary="Bulk import transactions from CSV/XLSX",
    description="Upload a CSV or Excel file and create transactions for the authenticated user.",
)
async def api_import_transactions(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: DB = Depends(get_db),
):
    filename = file.filename or ""
    content = await file.read()
    ext = filename.lower()
    transactions: List[dict] = []
    if ext.endswith(".xlsx") or file.content_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"):
        try:
            transactions = _parse_excel_transactions(content)
        except Exception as e:
            raise HTTPException(400, f"Failed to parse Excel file: {e}")
    elif ext.endswith(".csv") or "csv" in (file.content_type or ""):
        try:
            transactions = _parse_csv_transactions(content)
        except Exception as e:
            raise HTTPException(400, f"Failed to parse CSV file: {e}")
    else:
        raise HTTPException(400, "Unsupported file type. Please upload a .csv or .xlsx file.")

    created = 0
    errors: List[str] = []
    for txn in transactions:
        try:
            create_transaction(db, current_user["user_id"], txn)
            created += 1
        except Exception as e:
            errors.append(str(e))

    return {"imported": created, "skipped": len(transactions) - created, "errors": errors[:10]}


@router.get(
    "/calendar",
    summary="Calendar summary for a month",
    description="Returns per-day income/expense totals and transactions for the given month (YYYY-MM).",
)
def api_calendar_summary(month: str, current_user=Depends(get_current_user), db: DB = Depends(get_db)):
    if not month or len(month) != 7 or month[4] != "-":
        raise HTTPException(400, "month must be YYYY-MM")
    date_from = f"{month}-01"
    date_to = f"{month}-31"
    items = list_transactions(db, current_user["user_id"], date_from=date_from, date_to=date_to, category_id=None, uncategorized=False, limit=None)
    days: dict[str, dict] = {}
    for i in items:
        d = i.get("date")
        if not d:
            continue
        day = days.setdefault(d, {"date": d, "income": 0, "expense": 0, "transactions": []})
        amt = i.get("amount") or 0
        if amt >= 0:
            day["income"] += amt
        else:
            day["expense"] += abs(amt)
        day["transactions"].append({
            "txnId": i.get("txnId"),
            "date": i.get("date"),
            "merchant": i.get("merchant"),
            "description": i.get("description"),
            "amount": i.get("amount"),
            "currency": i.get("currency"),
            "categoryId": i.get("categoryId"),
            "notes": i.get("notes"),
            "source": i.get("source"),
            "accountId": i.get("accountId"),
            "receiptId": i.get("receiptId"),
        })
    # sort transactions by date/time if present
    summary = sorted(days.values(), key=lambda x: x["date"])
    return {"month": month, "days": summary}
